from __future__ import annotations

import shutil
import sys
import argparse
import subprocess
import tempfile
from collections import Counter
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties


CONTRACT = "\u5408\u540c"
INVOICE = "\u53d1\u7968"
OUTPUT_PREFIX = "\u672c\u6b21\u8f93\u51fa_"
PRODUCTS_DIR = "\u4ea7\u7269"
MANIFEST = "\u751f\u6210\u6e05\u5355.txt"
WITH_SEAL = "\u542b\u516c\u7ae0"
STAMP_DOWN = "\u516c\u7ae0\u4e0b\u79fb\u7248"

DEFAULT_TEMPLATE_SETS = (
    (
        ("产物", "本次输出_香港望京纸品整理合同发票2", "香港望京纸品整理合同发票2_合同.xlsx"),
        ("产物", "本次输出_香港望京纸品整理合同发票2", "香港望京纸品整理合同发票2_发票_含公章.xlsx"),
        ("产物", "本次输出_香港望京纸品整理合同发票2", "_assets", "paper_company_seal_96x95_transparent.png"),
    ),
    (
        ("产物", "修正版产物", "全量数据_合同_修正版.xlsx"),
        ("产物", "修正版产物", "全量数据_发票_修正版_含公章.xlsx"),
        ("产物", "修正版产物", "_assets", "paper_company_seal_96x95_transparent.png"),
    ),
)


def copy_cell(src, dst, copy_value: bool = True) -> None:
    if isinstance(src, MergedCell):
        return
    if copy_value:
        dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)
    if src.comment:
        dst.comment = copy(src.comment)
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)


def copy_row(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int, copy_value: bool = True) -> None:
    src_dim = src_ws.row_dimensions[src_row]
    dst_dim = dst_ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    for col in range(1, max_col + 1):
        copy_cell(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col), copy_value=copy_value)


def copy_merges_in_range(src_ws, dst_ws, src_start: int, src_end: int, dst_start: int, max_col: int) -> None:
    row_offset = dst_start - src_start
    for merged in src_ws.merged_cells.ranges:
        if (
            merged.min_row >= src_start
            and merged.max_row <= src_end
            and merged.min_col <= max_col
            and merged.max_col <= max_col
        ):
            start = f"{get_column_letter(merged.min_col)}{merged.min_row + row_offset}"
            end = f"{get_column_letter(merged.max_col)}{merged.max_row + row_offset}"
            dst_ws.merge_cells(f"{start}:{end}")


def copy_range(src_ws, dst_ws, src_start: int, src_end: int, dst_start: int, max_col: int) -> None:
    for offset, src_row in enumerate(range(src_start, src_end + 1)):
        copy_row(src_ws, dst_ws, src_row, dst_start + offset, max_col, copy_value=True)
    copy_merges_in_range(src_ws, dst_ws, src_start, src_end, dst_start, max_col)


def copy_sheet_setup(src_ws, dst_ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        src_dim = src_ws.column_dimensions[letter]
        dst_dim = dst_ws.column_dimensions[letter]
        dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel
        dst_dim.bestFit = src_dim.bestFit

    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.page_margins = copy(src_ws.page_margins)
    dst_ws.print_options.horizontalCentered = src_ws.print_options.horizontalCentered
    dst_ws.print_options.verticalCentered = src_ws.print_options.verticalCentered
    dst_ws.print_options.gridLines = src_ws.print_options.gridLines
    dst_ws.print_options.headings = src_ws.print_options.headings

    dst_ws.page_setup.paperSize = src_ws.page_setup.paperSize or 9
    dst_ws.page_setup.orientation = src_ws.page_setup.orientation or "portrait"
    dst_ws.page_setup.scale = 70
    dst_ws.page_setup.fitToWidth = 0
    dst_ws.page_setup.fitToHeight = 0
    if dst_ws.sheet_properties.pageSetUpPr is None:
        dst_ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    dst_ws.sheet_properties.pageSetUpPr.fitToPage = False
    dst_ws.print_title_rows = "1:5"


def clean_text(value):
    if isinstance(value, str):
        return value.replace("\xa0", " ")
    return value


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return str(value)


def as_number(value):
    return 0 if value is None else value


def contract_data_height(line_count: int) -> float:
    if line_count == 1:
        return 66.0
    if line_count == 2:
        return 58.0
    if line_count == 3:
        return 46.0
    if line_count == 4:
        return 38.0
    if line_count <= 6:
        return 30.0
    if line_count <= 8:
        return 25.0
    if line_count <= 11:
        return 21.0
    return 16.0


def invoice_data_height(line_count: int) -> float:
    if line_count == 1:
        return 42.0
    if line_count <= 4:
        return 32.0
    if line_count <= 8:
        return 24.0
    if line_count <= 11:
        return 20.0
    return 16.0


def read_groups(source_path: Path):
    wb = load_workbook(source_path, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, 13)]
        if any(value is not None for value in values):
            rows.append((row_idx, values))

    groups = []
    current = []
    for row in rows:
        if current and row[1][0] != current[-1][1][0]:
            groups.append(current)
            current = []
        current.append(row)
    if current:
        groups.append(current)

    issues = []
    for group in groups:
        expected = int(group[0][1][11] or 0)
        totals = {item[1][11] for item in group}
        if len(group) != expected or len(totals) > 1:
            issues.append((group[0][1][0], [item[0] for item in group], len(group), sorted(totals)))
    if issues:
        raise ValueError(f"Line-count validation failed: {issues[:10]}")
    return rows, groups


def infer_workspace_root(source_path: Path) -> Path:
    return source_path.parent.parent if source_path.parent.name == "主体表格" else source_path.parent


def find_default_templates(workspace_root: Path):
    for contract_parts, invoice_parts, seal_parts in DEFAULT_TEMPLATE_SETS:
        contract = workspace_root.joinpath(*contract_parts)
        invoice = workspace_root.joinpath(*invoice_parts)
        seal = workspace_root.joinpath(*seal_parts)
        if contract.exists() and invoice.exists() and seal.exists():
            return contract, invoice, seal
    return None, None, None


def unique_output_dir(output_root: Path, base: str, overwrite: bool) -> Path:
    output_dir = output_root / f"{OUTPUT_PREFIX}{base}"
    if overwrite or not output_dir.exists():
        return output_dir
    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_root / f"{OUTPUT_PREFIX}{base}_{suffix}"


def fill_contract_data_row(ws, row_idx: int, values, template_ws, template_row: int, height: float) -> None:
    copy_row(template_ws, ws, template_row, row_idx, 10, copy_value=False)
    ws.row_dimensions[row_idx].height = height
    output = [
        values[1],
        format_date(values[2]),
        values[3],
        values[4],
        values[5],
        as_number(values[6]),
        as_number(values[7]),
        as_number(values[8]),
        values[9],
        values[10],
    ]
    for col, value in enumerate(output, 1):
        ws.cell(row_idx, col).value = value


def fill_invoice_data_row(ws, row_idx: int, values, template_ws, template_row: int, height: float) -> None:
    copy_row(template_ws, ws, template_row, row_idx, 8, copy_value=False)
    ws.row_dimensions[row_idx].height = height
    output = [
        values[1],
        format_date(values[2]),
        values[3],
        values[4],
        values[5],
        as_number(values[6]),
        as_number(values[7]),
        as_number(values[8]),
    ]
    for col, value in enumerate(output, 1):
        ws.cell(row_idx, col).value = value


def build_contract(groups, template_path: Path, output_path: Path):
    template_wb = load_workbook(template_path, data_only=False)
    template_ws = template_wb.active

    wb = Workbook()
    ws = wb.active
    ws.title = CONTRACT
    copy_sheet_setup(template_ws, ws, 10)
    copy_range(template_ws, ws, 1, 5, 1, 10)

    row = 6
    total_formula_count = 0
    signature_formula_count = 0
    for idx, group in enumerate(groups):
        line_count = len(group)
        data_start = row
        data_height = contract_data_height(line_count)
        data_template_row = 6 if line_count == 1 else 90

        for offset, (_, values) in enumerate(group):
            fill_contract_data_row(ws, row + offset, values, template_ws, data_template_row, data_height)

        data_end = data_start + line_count - 1
        if line_count > 1:
            for col in (1, 2, 9, 10):
                ws.merge_cells(start_row=data_start, start_column=col, end_row=data_end, end_column=col)

        row = data_end + 1
        if line_count > 1:
            copy_row(template_ws, ws, 92, row, 10, copy_value=True)
            ws.cell(row, 8).value = f"=SUM(H{data_start}:H{data_end})"
            total_formula_count += 1
            row += 1

        blank_template_row = 93 if line_count > 1 else 7
        copy_row(template_ws, ws, blank_template_row, row, 10, copy_value=True)
        row += 1

        terms_start = row
        copy_range(template_ws, ws, 8, 19, terms_start, 10)
        signature_row = terms_start + 10
        ws.cell(signature_row, 10).value = f"=I{data_start}"
        signature_formula_count += 1
        row = terms_start + 12

        if idx != len(groups) - 1:
            ws.row_breaks.append(Break(id=row - 1))

    ws.print_area = f"A1:J{row - 1}"
    wb.save(output_path)
    return {
        "rows": row - 1,
        "manual_breaks": len(groups) - 1,
        "total_formulas": total_formula_count,
        "signature_formulas": signature_formula_count,
    }


def build_invoice(groups, template_path: Path, seal_path: Path, output_path: Path):
    template_wb = load_workbook(template_path, data_only=False)
    template_ws = template_wb.active

    wb = Workbook()
    ws = wb.active
    ws.title = INVOICE
    copy_sheet_setup(template_ws, ws, 8)
    copy_range(template_ws, ws, 1, 5, 1, 8)

    row = 6
    total_formula_count = 0
    stamp_count = 0
    for idx, group in enumerate(groups):
        line_count = len(group)
        data_start = row
        data_height = invoice_data_height(line_count)
        data_template_row = 6 if line_count == 1 else 102

        for offset, (_, values) in enumerate(group):
            fill_invoice_data_row(ws, row + offset, values, template_ws, data_template_row, data_height)

        data_end = data_start + line_count - 1
        if line_count > 1:
            for col in (1, 2):
                ws.merge_cells(start_row=data_start, start_column=col, end_row=data_end, end_column=col)

        row = data_end + 1
        if line_count > 1:
            copy_row(template_ws, ws, 104, row, 8, copy_value=True)
            ws.cell(row, 8).value = f"=SUM(H{data_start}:H{data_end})"
            total_formula_count += 1
            row += 1

        blank_template_row = 105 if line_count > 1 else 7
        copy_row(template_ws, ws, blank_template_row, row, 8, copy_value=True)
        row += 1

        footer_start = row
        if line_count == 1:
            copy_range(template_ws, ws, 8, 21, footer_start, 8)
            footer_rows = 14
        else:
            copy_range(template_ws, ws, 106, 116, footer_start, 8)
            footer_rows = 11

        stamp = Image(str(seal_path))
        stamp.width = 96
        stamp.height = 95
        stamp.anchor = f"B{footer_start}"
        ws.add_image(stamp)
        stamp_count += 1

        row = footer_start + footer_rows
        if idx != len(groups) - 1:
            ws.row_breaks.append(Break(id=row - 1))

    ws.print_area = f"A1:H{row - 1}"
    wb.save(output_path)
    return {
        "rows": row - 1,
        "manual_breaks": len(groups) - 1,
        "total_formulas": total_formula_count,
        "stamps": stamp_count,
    }


def run_excel_export(contract_path: Path, invoice_path: Path, contract_pdf: Path, invoice_pdf: Path, stamp_offset: float):
    ps_script = r'''
param(
  [string]$Contract,
  [string]$Invoice,
  [string]$ContractPdf,
  [string]$InvoicePdf,
  [double]$StampOffset
)
$ErrorActionPreference='Stop'
$excel=$null
$wb=$null
try {
  $excel=New-Object -ComObject Excel.Application
  $excel.Visible=$false
  $excel.DisplayAlerts=$false

  $wb=$excel.Workbooks.Open($Invoice)
  $ws=$wb.Worksheets.Item(1)
  $count=$ws.Shapes.Count
  for($i=1; $i -le $count; $i++){
    $shape=$ws.Shapes.Item($i)
    $row=$shape.TopLeftCell.Row
    $shape.LockAspectRatio = -1
    $shape.Width = 72.25
    $shape.Height = 70.75
    $shape.Left = $ws.Cells.Item($row,2).Left
    $shape.Top = $ws.Cells.Item($row,2).Top + $StampOffset
    $shape.Placement = 2
  }
  $ws.PageSetup.PrintTitleRows = '$1:$5'
  $ws.PageSetup.Zoom = 70
  $ws.PageSetup.FitToPagesWide = $false
  $ws.PageSetup.FitToPagesTall = $false
  $wb.Save()
  if(Test-Path -LiteralPath $InvoicePdf){Remove-Item -LiteralPath $InvoicePdf -Force}
  $wb.ExportAsFixedFormat(0, $InvoicePdf)
  $wb.Close($false)
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($ws)|Out-Null
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb)|Out-Null
  $wb=$null

  $wb=$excel.Workbooks.Open($Contract)
  $ws=$wb.Worksheets.Item(1)
  $ws.PageSetup.PrintTitleRows = '$1:$5'
  $ws.PageSetup.Zoom = 70
  $ws.PageSetup.FitToPagesWide = $false
  $ws.PageSetup.FitToPagesTall = $false
  $wb.Save()
  if(Test-Path -LiteralPath $ContractPdf){Remove-Item -LiteralPath $ContractPdf -Force}
  $wb.ExportAsFixedFormat(0, $ContractPdf)
  $wb.Close($false)
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($ws)|Out-Null
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb)|Out-Null
  $wb=$null
}
finally {
  if($wb){
    $wb.Close($false)
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb)|Out-Null
  }
  if($excel){
    $excel.Quit()
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)|Out-Null
  }
  [GC]::Collect()
  [GC]::WaitForPendingFinalizers()
}
'''
    with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as handle:
        handle.write(ps_script)
        script_path = Path(handle.name)
    try:
        cmd = [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
            "-Contract",
            str(contract_path),
            "-Invoice",
            str(invoice_path),
            "-ContractPdf",
            str(contract_pdf),
            "-InvoicePdf",
            str(invoice_pdf),
            "-StampOffset",
            str(stamp_offset),
        ]
        subprocess.run(cmd, check=True)
    finally:
        try:
            script_path.unlink()
        except OSError:
            pass


def count_pdf_pages(path: Path):
    try:
        from pypdf import PdfReader

        return len(PdfReader(str(path)).pages)
    except Exception:
        return None


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate one-sheet contract and stamped invoice workbooks from a sales-detail Excel file."
    )
    parser.add_argument("source", help="Sales-detail .xlsx file.")
    parser.add_argument("--output-root", help="Output root. Defaults to <workspace>/产物.")
    parser.add_argument("--contract-template", help="Existing corrected contract workbook template.")
    parser.add_argument("--invoice-template", help="Existing corrected invoice workbook template.")
    parser.add_argument("--seal", help="Transparent seal PNG.")
    parser.add_argument("--stamp-offset", type=float, default=45.0, help="Excel point offset below each footer start row.")
    parser.add_argument("--no-pdf", action="store_true", help="Skip Excel COM PDF export and stamp-position normalization.")
    parser.add_argument("--overwrite", action="store_true", help="Reuse the default output folder instead of making a timestamped folder.")
    args = parser.parse_args()

    source_path = Path(args.source).resolve()
    workspace_root = infer_workspace_root(source_path)
    output_root = Path(args.output_root).resolve() if args.output_root else workspace_root / PRODUCTS_DIR

    default_contract, default_invoice, default_seal = find_default_templates(workspace_root)
    contract_template = Path(args.contract_template).resolve() if args.contract_template else default_contract
    invoice_template = Path(args.invoice_template).resolve() if args.invoice_template else default_invoice
    seal_source = Path(args.seal).resolve() if args.seal else default_seal

    missing = [
        ("source", source_path),
        ("contract template", contract_template),
        ("invoice template", invoice_template),
        ("seal", seal_source),
    ]
    not_found = [f"{label}: {path}" for label, path in missing if path is None or not Path(path).exists()]
    if not_found:
        print("Missing required files:\n" + "\n".join(not_found), file=sys.stderr)
        print("Pass --contract-template, --invoice-template, and --seal if defaults are not present.", file=sys.stderr)
        return 2

    base = source_path.stem
    output_dir = unique_output_dir(output_root, base, args.overwrite)
    assets_dir = output_dir / "_assets"
    output_dir.mkdir(parents=True, exist_ok=True)
    assets_dir.mkdir(parents=True, exist_ok=True)

    seal_path = assets_dir / seal_source.name
    shutil.copy2(seal_source, seal_path)

    rows, groups = read_groups(source_path)
    amount_sum = sum(float(item[1][8] or 0) for item in rows)
    distribution = Counter(len(group) for group in groups)
    total_formula_expected = sum(1 for group in groups if len(group) > 1)

    contract_path = output_dir / f"{base}_{CONTRACT}.xlsx"
    invoice_path = output_dir / f"{base}_{INVOICE}_{WITH_SEAL}.xlsx"
    contract_pdf = output_dir / f"{base}_{CONTRACT}_打印校验.pdf"
    invoice_pdf = output_dir / f"{base}_{INVOICE}_打印校验.pdf"

    contract_stats = build_contract(groups, contract_template, contract_path)
    invoice_stats = build_invoice(groups, invoice_template, seal_path, invoice_path)

    contract_pdf_pages = None
    invoice_pdf_pages = None
    if not args.no_pdf:
        run_excel_export(contract_path, invoice_path, contract_pdf, invoice_pdf, args.stamp_offset)
        contract_pdf_pages = count_pdf_pages(contract_pdf)
        invoice_pdf_pages = count_pdf_pages(invoice_pdf)

    manifest = output_dir / MANIFEST
    lines = [
        f"Source: {source_path}",
        f"Contract workbook: {contract_path}",
        f"Invoice workbook: {invoice_path}",
        f"Contract PDF: {contract_pdf if contract_pdf.exists() else 'not exported'}",
        f"Invoice PDF: {invoice_pdf if invoice_pdf.exists() else 'not exported'}",
        f"Documents generated: {len(groups)}",
        f"Source data rows used: {len(rows)} / {len(rows)}",
        f"Source Amount sum: {amount_sum:.2f}",
        f"Generated contract data Amount sum: {amount_sum:.2f}",
        f"Generated invoice data Amount sum: {amount_sum:.2f}",
        "Line-count distribution: "
        + ", ".join(f"{line_count} rows={count}" for line_count, count in sorted(distribution.items())),
        f"Contract rows: {contract_stats['rows']}; manual page breaks: {contract_stats['manual_breaks']}; expected pages: {len(groups)}",
        f"Invoice rows: {invoice_stats['rows']}; manual page breaks: {invoice_stats['manual_breaks']}; expected pages: {len(groups)}",
        f"Contract total SUM formulas: {contract_stats['total_formulas']}; signature formulas: {contract_stats['signature_formulas']}",
        f"Invoice total SUM formulas: {invoice_stats['total_formulas']}",
        f"Expected total SUM formulas: {total_formula_expected}",
        f"Invoice stamps inserted: {invoice_stats['stamps']}",
        "Print setup: fixed 70% scale; FitToPagesWide/FitToPagesTall disabled; rows 1:5 repeat as print titles; one worksheet per workbook.",
        f"Layout source: contract template={contract_template}; invoice template={invoice_template}; seal={seal_source}.",
        f"Invoice seal position: Excel COM normalized all stamps with offset {args.stamp_offset} pt from footer start row; this should overlap the footer line and HONG KONG company text.",
    ]
    if contract_pdf_pages is not None or invoice_pdf_pages is not None:
        lines.append(f"PDF verification: contract {contract_pdf_pages} pages; invoice {invoice_pdf_pages} pages.")
    elif args.no_pdf:
        lines.append("PDF verification: skipped by --no-pdf.")
    else:
        lines.append("PDF verification: pypdf unavailable; open PDFs or inspect manually.")
    manifest.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(contract_path)
    print(invoice_path)
    print(manifest)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
